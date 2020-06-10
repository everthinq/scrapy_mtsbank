import json
import scrapy
from openpyxl import load_workbook


class mtsbank(scrapy.Spider):
    name = 'mtsbank'

    def start_requests(self):
        region_codes = [
            {"region_name": "Армавир", "city": "fc9c55d0-c66e-455e-8034-b0944b025c38"},
            {"region_name": "Амурск", "city": "d3c4b43d-3e19-4454-939b-d92ef3d6c875"},
            {"region_name": "Ангарск", "city": "82b6b7c8-82a4-44b2-8bc7-691373706b89"},
            {"region_name": "Бикин", "city": "70dc89f8-f067-482a-bd50-4fb2782463e7"},
            {"region_name": "Благовещенск", "city": "8f41253d-6e3b-48a9-842a-25ba894bd093"},
            {"region_name": "Белогорск", "city": "c528e99b-7e81-4290-9cda-8713884472a5"},
            {"region_name": "Биробиджан", "city": "5d133391-46ee-496b-83a6-efeeaa903643"},
            {"region_name": "Владивосток", "city": "7b6de6a5-86d0-4735-b11a-499081111af8"},
            {"region_name": "Ванино", "city": "2a148048-2ee1-4009-8c28-7de5b03afa35"},
            {"region_name": "Вяземский", "city": "6dbd1c8b-7c73-4ff7-9cbd-7062ac0cc88d"},
            {"region_name": "Волгоград", "city": "a52b7389-0cfe-46fb-ae15-298652a64cf8"},
            {"region_name": "Волжский", "city": "bc5ed788-84c8-493e-9598-7a15a9f1e4c1"},
            {"region_name": "Вологда", "city": "023484a5-f98d-4849-82e1-b7e0444b54ef"},
            {"region_name": "Воронеж", "city": "5bf5ddff-6353-4a3d-80c4-6fb27f00c6c1"},
            {"region_name": "Де-Кастри", "city": "41b5a65a-51cd-4602-8e09-9ecae29c96fc"},
            {"region_name": "Екатеринбург", "city": "2763c110-cb8b-416a-9dac-ad28a55b4402"},
            {"region_name": "Иркутск", "city": "8eeed222-72e7-47c3-ab3a-9a553c31cf72"},
            {"region_name": "Казань", "city": "93b3df57-4c89-44df-ac42-96f05e9cd3b9"},
            {"region_name": "Краснодар", "city": "7dfa745e-aa19-4688-b121-b655c11e482f"},
            {"region_name": "Красноярск", "city": "9b968c73-f4d4-4012-8da8-3dacd4d4c1bd"},
            {"region_name": "Комсомольск-на-Амуре", "city": "a29c5b20-5056-412b-9af6-7b805aa3ea72"},
            {"region_name": "Калининград", "city": "df679694-d505-4dd3-b514-4ba48c8a97d8"},
            {"region_name": "Киров", "city": "452a2ddf-88a1-4e35-8d8d-8635493768d4"},
            {"region_name": "Москва", "city": "77"},
            {"region_name": "Нефтекамск", "city": "2c9997d2-ce94-431a-96c9-722d2238d5c8"},
            {"region_name": "Новороссийск", "city": "16ac039a-5257-4715-a8c5-d6bd9e617b53"},
            {"region_name": "Находка", "city": "225a3506-35aa-4456-8bd7-244bdfbc4eaf"},
            {"region_name": "Николаевск-на-Амуре", "city": "7a58fb7c-6d03-46e4-b5fc-3f5b587c09be"},
            {"region_name": "Нижний Новгород", "city": "555e7d61-d9a7-4ba6-9770-6caa8198c483"},
            {"region_name": "Новосибирск", "city": "8dea00e3-9aab-4d8e-887c-ef2aaa546456"},
            {"region_name": "Нижний Тагил", "city": "cc73d6af-6e2e-4a1f-be8e-682c289b0b57"},
            {"region_name": "Октябрьский", "city": "abd1bc35-ec51-437a-abee-76a4f620f662"},
            {"region_name": "Омск", "city": "140e31da-27bf-4519-9ea0-6185d681d44e"},
            {"region_name": "Петрозаводск", "city": "ccc34487-8fd4-4e71-b032-f4e6c82fb354"},
            {"region_name": "Переяславка", "city": "826fad9e-dbb9-454c-8890-206892c890bc"},
            {"region_name": "Пермь", "city": "a309e4ce-2f36-4106-b1ca-53e0f48a6d95"},
            {"region_name": "Ростов-на-Дону", "city": "c1cfe4b9-f7c2-423c-abfa-6ed1c05a15c5"},
            {"region_name": "Рязань", "city": "86e5bae4-ef58-4031-b34f-5e9ff914cd55"},
            {"region_name": "Санкт-Петербург", "city": "78"},
            {"region_name": "Стерлитамак", "city": "84e0b23d-82fe-40a8-8739-55e679780dc3"},
            {"region_name": "Сегежа", "city": "4eb67866-2460-40c3-b69f-80889385caa3"},
            {"region_name": "Сыктывкар", "city": "d2944a73-daf4-4a08-9b34-d9b0af7785a1"},
            {"region_name": "Сочи", "city": "79da737a-603b-4c19-9b54-9114c96fb912"},
            {"region_name": "Ставрополь", "city": "2a1c7bdb-05ea-492f-9e1c-b3999f79dcbc"},
            {"region_name": "Советская Гавань", "city": "64f0ebe7-00e4-40a0-9dd9-15f9293632ae"},
            {"region_name": "Солнечный", "city": "2f1bc221-8ccc-4f89-8a24-0dea035b0a9c"},
            {"region_name": "Соловьевск", "city": "ef16d17b-5db2-4d22-bf7a-fcc8c5349b12"},
            {"region_name": "Самара", "city": "bb035cc3-1dc2-4627-9d25-a1bf2d4b936b"},
            {"region_name": "Саратов", "city": "bf465fda-7834-47d5-986b-ccdb584a85a6"},
            {"region_name": "Туймазы", "city": "511a0136-f60c-451b-a2eb-3402103f1223"},
            {"region_name": "Тында", "city": "007e010f-e110-4a55-90a7-c4acac623c9b"},
            {"region_name": "Томск", "city": "e3b0eae8-a4ce-4779-ae04-5c0797de66be"},
            {"region_name": "Тюмень", "city": "9ae64229-9f7b-4149-b27a-d1f6ec74b5ce"},
            {"region_name": "Уфа", "city": "7339e834-2cb4-4734-a4c7-1fca2c66e562"},
            {"region_name": "Ухта", "city": "067b4cef-e128-4d5a-8305-fecf53e7b7e8"},
            {"region_name": "Уссурийск", "city": "de7335fb-9baa-48eb-927d-0bb299b2e5bc"},
            {"region_name": "Хабаровск", "city": "a4859da8-9977-4b62-8436-4e1b98c5d13f"},
            {"region_name": "Хор", "city": "871f14e0-b731-416b-b294-a9caeca3a464"},
            {"region_name": "Чегдомын", "city": "92fa4d67-876c-4a2f-a6db-44bf30124b67"},
            {"region_name": "Челябинск", "city": "a376e68d-724a-4472-be7c-891bdb09ae32"},
            {"region_name": "Чита", "city": "2d9abaa6-85a6-4f1f-a1bd-14b76ec17d9c"},
            {"region_name": "Южно-Сахалинск", "city": "44388ad0-06aa-49b0-bbf9-1704629d1d68"}
        ]
        self.cell_value = '2'
        self.workbook = load_workbook('51. PJSC MTS Bank Russia.xlsx')
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

        for region in region_codes:
            city = region['city']
            region_name = region['region_name']

            yield scrapy.Request(
                method='GET',
                url='https://www.mtsbank.ru/ajax/offices_v2.php?city=' + city + '&region_name=' + region_name,
                callback=self.parse_json
            )

    def parse_json(self, response):
        if (response.text):
            JSON = json.loads(response.text)

            for branch in JSON['data']:
                branch_name = branch['name']
                address = branch['address']['full']
                lat = branch['address']['coordinates'][0]
                lng = branch['address']['coordinates'][1]

                print('Writing --', 'name:', branch_name, 'address:',
                      address, 'lat:', lat, 'lng:', lng)

                self.cell_value = str(self.cell_value)
                self.worksheet['B' + self.cell_value] = 'PJSC MTS Bank Russia'
                self.worksheet['C' + self.cell_value] = branch_name
                self.worksheet['D' + self.cell_value] = address
                self.worksheet['G' + self.cell_value] = 'Russia'
                self.worksheet['H' + self.cell_value] = 'RU'
                self.worksheet['M' + self.cell_value] = lat
                self.worksheet['N' + self.cell_value] = lng
                self.worksheet['O' + self.cell_value] = 'Address'
                self.worksheet['R' + self.cell_value] = 'Bank website'
                self.cell_value = int(self.cell_value) + 1

        self.workbook.save('51. PJSC MTS Bank Russia.xlsx')
